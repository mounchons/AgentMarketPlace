#!/usr/bin/env python3
"""
export_xlsx.py — render a quotation Markdown draft into a formatted .xlsx.

Usage:
    python export_xlsx.py <quotation.md> [output.xlsx]

Parses the quotation .md contract (see skills/quotation-builder/references/format-spec.md):
  - YAML frontmatter  -> header metadata (title, tech_stack, duration, warranty, notes, manday_rate)
  - the scope table   -> the first table whose header has both "Module" and "Price"
  - the out-of-scope table -> the first table whose header has "Manday"

Never invents a price. Blank Price cells stay blank; it warns if any are still blank but still exports.

Requires: PyYAML, openpyxl  (pip install pyyaml openpyxl)
"""
import sys
import re
import os

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


def die(msg, code=1):
    print("ERROR: " + msg, file=sys.stderr)
    sys.exit(code)


try:
    import yaml
except ImportError:
    die("PyYAML is not installed. Run: pip install pyyaml")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    die("openpyxl is not installed. Run: pip install openpyxl")


# ----------------------------------------------------------------------------- parsing helpers
def split_frontmatter(text):
    """Return (meta_dict, body_text). Frontmatter is a leading --- ... --- block."""
    if text.startswith("﻿"):
        text = text.lstrip("﻿")
    lines = text.splitlines()
    if not lines or lines[0].strip() != "---":
        return {}, text
    end = None
    for i in range(1, len(lines)):
        if lines[i].strip() == "---":
            end = i
            break
    if end is None:
        return {}, text
    fm = "\n".join(lines[1:end])
    body = "\n".join(lines[end + 1:])
    try:
        meta = yaml.safe_load(fm) or {}
    except yaml.YAMLError as e:
        die("frontmatter is not valid YAML: %s" % e)
    if not isinstance(meta, dict):
        die("frontmatter did not parse to a mapping")
    return meta, body


def split_row(line):
    return [c.strip() for c in line.strip().strip("|").split("|")]


def is_separator(line):
    body = line.strip().strip("|").replace("|", "").replace(":", "").replace("-", "").replace(" ", "")
    return body == "" and "-" in line


def extract_tables(text):
    """Find every Markdown pipe-table. Returns list of (header:list[str], rows:list[list[str]])."""
    lines = text.splitlines()
    tables = []
    i, n = 0, len(lines)
    while i < n:
        line = lines[i].strip()
        if line.startswith("|") and i + 1 < n and is_separator(lines[i + 1]):
            header = split_row(lines[i])
            rows = []
            j = i + 2
            while j < n and lines[j].strip().startswith("|"):
                if not is_separator(lines[j]):
                    rows.append(split_row(lines[j]))
                j += 1
            tables.append((header, rows))
            i = j
        else:
            i += 1
    return tables


def col_index(header, *keywords):
    """First column index whose header text contains any keyword (case-insensitive)."""
    for idx, h in enumerate(header):
        low = h.lower()
        for kw in keywords:
            if kw.lower() in low:
                return idx
    return None


def cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return row[idx].strip()


def strip_bold(s):
    s = s.strip()
    while s.startswith("*"):
        s = s[1:]
    while s.endswith("*"):
        s = s[:-1]
    return s.strip()


def parse_amount(s):
    """Parse a Price cell to a number, or None if blank/unparseable."""
    if s is None:
        return None
    t = str(s).strip()
    if t in ("", "-", "—"):
        return None
    t = t.replace(",", "").replace("฿", "").replace("บาท", "").replace(" ", "")
    m = re.search(r"-?\d+(\.\d+)?", t)
    if not m:
        return None
    val = float(m.group())
    return int(val) if val.is_integer() else val


# ----------------------------------------------------------------------------- data model
def parse_scope(table):
    """table = (header, rows). Returns (modules, total_price, blank_price_count)."""
    header, rows = table
    i_no = col_index(header, "no", "no.", "ลำดับ")
    i_mod = col_index(header, "module", "โมดูล")
    i_desc = col_index(header, "description", "รายละเอียด")
    i_comment = col_index(header, "comment", "หมายเหตุ")
    i_price = col_index(header, "price", "ราคา")
    if i_mod is None or i_price is None:
        die("scope table must have a Module column and a Price column")

    modules = []
    total_price = None
    blank_prices = 0
    current = None
    for row in rows:
        mod_raw = cell(row, i_mod)
        mod_txt = strip_bold(mod_raw)
        desc = cell(row, i_desc)
        comment = cell(row, i_comment)
        no = cell(row, i_no)
        price_raw = cell(row, i_price)
        price = parse_amount(price_raw)

        if mod_txt and re.search(r"(รวมทั้งหมด|total)", mod_txt, re.IGNORECASE):
            total_price = price
            continue
        if mod_txt:  # module header row
            current = {"no": no, "name": mod_txt, "price": price, "items": []}
            if price is None:
                blank_prices += 1
            modules.append(current)
        elif desc:  # line item
            if current is None:
                current = {"no": "", "name": "", "price": None, "items": []}
                modules.append(current)
            current["items"].append({"desc": desc, "comment": comment})
    return modules, total_price, blank_prices


def parse_oos(table):
    header, rows = table
    i_no = col_index(header, "no", "ลำดับ")
    i_item = col_index(header, "รายการ", "item")
    i_note = col_index(header, "หมายเหตุ", "note", "comment")
    i_manday = col_index(header, "manday", "แมนเดย์")
    out = []
    for row in rows:
        item = cell(row, i_item) if i_item is not None else (cell(row, 1) if len(row) > 1 else "")
        if not strip_bold(item):
            continue
        out.append({
            "no": cell(row, i_no),
            "item": strip_bold(item),
            "note": cell(row, i_note),
            "manday": cell(row, i_manday),
        })
    return out


# ----------------------------------------------------------------------------- workbook build
FONT_NAME = "Tahoma"
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="305496")
MODULE_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
OOS_FILL = PatternFill("solid", fgColor="EDEDED")


def build_workbook(meta, modules, total_price, oos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    widths = {"A": 6, "B": 42, "C": 58, "D": 34, "E": 15}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    currency = str(meta.get("currency") or "THB")
    r = 1

    def merge_line(text, bold=False, size=10, center=False, fill=None):
        nonlocal r
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name=FONT_NAME, size=size, bold=bold)
        c.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        if fill:
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = fill
        r += 1

    # --- title
    merge_line(str(meta.get("title") or "ใบเสนอราคา"), bold=True, size=14, center=True)
    for extra in ("client", "quote_no", "date"):
        val = meta.get(extra)
        if val:
            label = {"client": "ลูกค้า", "quote_no": "เลขที่", "date": "วันที่"}[extra]
            merge_line("%s: %s" % (label, val), size=10)
    r += 1

    # --- requirements
    tech = meta.get("tech_stack") or []
    if tech:
        merge_line("ความต้องการของระบบ (System Requirements)", bold=True)
        for line in tech:
            merge_line("• " + str(line))
    if meta.get("duration"):
        merge_line("ระยะเวลาพัฒนา: %s" % meta["duration"], bold=True)
    if meta.get("warranty"):
        merge_line("รับประกันผลงาน: %s" % meta["warranty"], bold=True)

    notes = meta.get("notes") or []
    if notes:
        r += 1
        merge_line("หมายเหตุ", bold=True)
        for note in notes:
            merge_line("• " + str(note))
    r += 1

    # --- scope table header
    headers = ["No", "Module", "Description", "Comment", "Price (%s)" % currency]
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=r, column=col, value=text)
        c.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    r += 1

    def styled(row, col, value, bold=False, wrap=True, align="left", fill=None, money=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name=FONT_NAME, size=10, bold=bold)
        c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
        c.border = BORDER
        if fill:
            c.fill = fill
        if money and isinstance(value, (int, float)):
            c.number_format = "#,##0"
        return c

    # --- scope rows
    for mod in modules:
        styled(r, 1, mod["no"], bold=True, align="center", fill=MODULE_FILL)
        styled(r, 2, mod["name"], bold=True, fill=MODULE_FILL)
        styled(r, 3, "", fill=MODULE_FILL)
        styled(r, 4, "", fill=MODULE_FILL)
        styled(r, 5, mod["price"] if mod["price"] is not None else "",
               bold=True, align="right", fill=MODULE_FILL, money=True)
        r += 1
        for item in mod["items"]:
            styled(r, 1, "")
            styled(r, 2, "")
            styled(r, 3, item["desc"])
            styled(r, 4, item["comment"])
            styled(r, 5, "", align="right")
            r += 1

    # --- total row
    styled(r, 1, "", fill=TOTAL_FILL)
    styled(r, 2, "รวมทั้งหมด (Total)", bold=True, fill=TOTAL_FILL)
    styled(r, 3, "", fill=TOTAL_FILL)
    styled(r, 4, "", fill=TOTAL_FILL)
    styled(r, 5, total_price if total_price is not None else "",
           bold=True, align="right", fill=TOTAL_FILL, money=True)
    r += 2

    # --- out of scope
    if oos:
        merge_line("งานนอกขอบเขต (Out of Scope / Optional)", bold=True, size=11, fill=OOS_FILL)
        oos_headers = ["No", "รายการ", "หมายเหตุ", "Manday"]
        # use columns 1..4 for the oos table
        for col, text in enumerate(oos_headers, start=1):
            c = ws.cell(row=r, column=col, value=text)
            c.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        r += 1
        for idx, row in enumerate(oos, start=1):
            styled(r, 1, row["no"] or idx, align="center")
            styled(r, 2, row["item"])
            styled(r, 3, row["note"])
            styled(r, 4, row["manday"], align="center")
            r += 1
        r += 1
        rate = meta.get("manday_rate")
        rate_txt = ("{:,}".format(rate) if isinstance(rate, (int, float)) else "______")
        merge_line("อัตราค่าบริการ: 1 Manday = %s บาท" % rate_txt, bold=True)

    return wb


# ----------------------------------------------------------------------------- main
def main():
    if len(sys.argv) < 2:
        die("usage: python export_xlsx.py <quotation.md> [output.xlsx]")
    md_path = sys.argv[1]
    if not os.path.isfile(md_path):
        die("input file not found: %s" % md_path)
    out_path = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(md_path)[0] + ".xlsx"

    with open(md_path, "r", encoding="utf-8") as f:
        text = f.read()

    meta, body = split_frontmatter(text)
    if str(meta.get("doc_type", "")).strip() and meta.get("doc_type") != "quotation":
        print("WARNING: doc_type is '%s', expected 'quotation'" % meta.get("doc_type"))

    tables = extract_tables(text)
    scope_table = next(((h, rws) for (h, rws) in tables
                        if col_index(h, "module") is not None and col_index(h, "price") is not None), None)
    if scope_table is None:
        die("no scope table found (need a table with Module and Price columns)")
    oos_table = next(((h, rws) for (h, rws) in tables if col_index(h, "manday") is not None), None)

    modules, total_price, blank_prices = parse_scope(scope_table)
    oos = parse_oos(oos_table) if oos_table else []

    wb = build_workbook(meta, modules, total_price, oos)
    wb.save(out_path)

    n_items = sum(len(m["items"]) for m in modules)
    print("OK  ->  %s" % out_path)
    print("    modules=%d  line_items=%d  out_of_scope=%d" % (len(modules), n_items, len(oos)))
    if total_price is None and blank_prices == len(modules):
        print("    NOTE: all prices are blank (draft). Fill the Price column, then re-export.")
    elif blank_prices or total_price is None:
        missing = []
        if blank_prices:
            missing.append("%d module price(s)" % blank_prices)
        if total_price is None:
            missing.append("grand total")
        print("    NOTE: still blank -> " + ", ".join(missing))


if __name__ == "__main__":
    main()
